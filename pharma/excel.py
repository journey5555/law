"""지출결의서 Excel 생성 — 공통/합계 셀 병합, 품목만 행 반복"""
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_THIN = Side(style="thin")

def _border():
    return Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)

def _hfill():
    return PatternFill("solid", fgColor="D9E1F2")

def _bold():
    return Font(bold=True, size=10)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


HEADERS = [
    "문서번호", "작성일자", "작성부서", "계약번호", "프로젝트명",
    "업체명", "사업자등록번호", "대표자", "연락처", "계좌번호",
    "지급방법", "지급예정일",
    "품목명", "수량", "단가", "금액", "품목비고",
    "합계금액", "합계(한글)",
    "적요", "비고",
]

COL_WIDTHS = [14, 12, 12, 14, 30, 18, 16, 10, 16, 24,
              10, 12, 24, 8, 12, 12, 16, 12, 24, 30, 20]

# 품목별로 반복되는 열 인덱스 (1-based): 13~17
_ITEM_COLS  = set(range(13, 18))
# 셀 병합 대상 열: 공통(1-12) + 합계/적요/비고(18-21)
_MERGE_COLS = set(range(1, 13)) | set(range(18, 22))
# 가운데 정렬 열
_CENTER_COLS = {1, 2, 3, 4, 11, 12, 14, 15, 16}
# 금액 형식 열 (단가=15, 품목금액=16, 합계금액=18)
_MONEY_COLS = {15, 16, 18}


def _to_number(val):
    """문자열 금액("1,000,000" 또는 "1000000") → int, 변환 불가 시 원본 반환"""
    if val is None:
        return val
    try:
        return int(str(val).replace(",", "").replace(" ", ""))
    except (ValueError, TypeError):
        return val


def generate(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "지출결의서"

    from openpyxl.utils import get_column_letter
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 헤더 행
    for col, header in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font      = _bold()
        c.fill      = _hfill()
        c.alignment = _center()
        c.border    = _border()
    ws.row_dimensions[1].height = 22

    v     = data.get("vendor")  or {}
    p     = data.get("payment") or {}
    t     = data.get("total")   or {}
    items = data.get("items")   or [{}]

    common = [
        data.get("document_id"),
        data.get("created_date"),
        data.get("department"),
        data.get("contract_no"),
        data.get("project_name"),
        v.get("company_name"),
        v.get("business_registration_no"),
        v.get("representative"),
        v.get("contact"),
        v.get("account_no"),
        p.get("method"),
        p.get("scheduled_date"),
    ]
    tail = [
        t.get("amount"),
        t.get("amount_korean"),
        data.get("description"),
        data.get("remark"),
    ]

    first_data_row = 2
    last_data_row  = first_data_row + len(items) - 1

    # 품목 행 작성
    for row_idx, item in enumerate(items, first_data_row):
        row_vals = common + [
            item.get("name"),
            item.get("quantity"),
            item.get("unit_price"),
            item.get("amount"),
            item.get("remark"),
        ] + tail

        for col, val in enumerate(row_vals, 1):
            # 병합 대상은 첫 행에만 값 입력
            if col in _MERGE_COLS and row_idx != first_data_row:
                val = None
            if col in _MONEY_COLS:
                val = _to_number(val)
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = _center() if col in _CENTER_COLS else _left()
            c.border    = _border()
            if col in _MONEY_COLS:
                c.number_format = "#,##0"
        ws.row_dimensions[row_idx].height = 18

    # 셀 병합 (품목이 2개 이상일 때만)
    if len(items) > 1:
        for col in _MERGE_COLS:
            ws.merge_cells(
                start_row=first_data_row, start_column=col,
                end_row=last_data_row,   end_column=col,
            )
            # 병합된 셀 스타일 재지정 (merge 후 top-left 셀에만 스타일 남음)
            c = ws.cell(row=first_data_row, column=col)
            c.alignment = _center() if col in _CENTER_COLS else _left()
            c.border    = _border()
            if col in _MONEY_COLS:
                c.number_format = "#,##0"

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
